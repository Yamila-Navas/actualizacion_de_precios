import openpyxl as o
from .models import Producto

def procesar_excel(archivo):
    # Abre el archivo exel
    archivo_excel = o.load_workbook(archivo)

    # Obtiene hoja activa
    hoja = archivo_excel.active

    # Lista para almacenar las filas del Excel
    filasExel = ['referencia', 'codigo', 'nombre', 'precio']

    # Recorrer fila por fila
    for fila in hoja.iter_rows(values_only=True):
        ls = []
        for celda in fila:
            
            if celda is not None:
                ls.append(celda)

        if len(ls) == 4 and ls[0] != "Ref.":
            filasExel.append(ls)

    # Crea un nuevo libro de trabajo (archivo Excel)
    nuevoExel = o.Workbook()

    # Obtiene la hoja activa (hoja por defecto)
    hoja = nuevoExel.active

    # Escribe los datos en el archivo Excel
    for fila, fila_data in enumerate(filasExel):
        for columna, celda_data in enumerate(fila_data):
            hoja.cell(row=fila + 1, column=columna + 1, value=celda_data)

    # Guarda el archivo Excel
    nuevoExel.save('nuevo_archivo.xlsx')

    actualizar_base_de_datos('nuevo_archivo.xlsx')
    








def actualizar_base_de_datos(archivo_excel):
    # Abre el archivo Excel
    workbook = o.load_workbook(archivo_excel)
    sheet = workbook.active

    # Itera sobre las filas del archivo Excel
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Suponiendo que la primera fila es encabezado
        codigo = row[1]  # Supongamos que la columna "código" está en la primera posición
        nuevo_precio = row[3]  # Supongamos que el nuevo precio está en la segunda columna

        try:
            # Intenta buscar el objeto en la base de datos con el código correspondiente
            producto = Producto.objects.get(codigo=codigo)

            # Actualiza el precio del producto en la base de datos
            producto.precio = nuevo_precio
            producto.save()
        except Producto.DoesNotExist:
            # No se encontró un producto con ese código, simplemente continúa con la siguiente fila
            pass
